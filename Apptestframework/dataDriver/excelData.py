#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2021/3/13 9:08
# @Author : ZY
# @File : excelData.py
# @Project : APP

# from openpyxl import load_workbook
# from log_ri_zhi.logHandler import LogHandler
#
#
# # 读
# class ReadExcel():
#     logger = LogHandler.logger
#
#     def __init__(self, path):
#         self.workbook = load_workbook(path)
#
#     # 获取标签页
#     def get_sheetname(self):
#         return self.workbook.sheetnames
#
#     # def isTrue(self, c, num=0):
#     #     if num > 0:
#     #         pass
#     #     elif num > c:
#
#     # 读取所有行
#     # def get_all_rows(self, sheetname, num=0):
#     #     try:
#     #         worklist = []
#     #         sheet = self.workbook[sheetname]
#     #         for i in sheet.rows:
#     #             rowlist = []
#     #             for j in i:
#     #                 rowlist.append(j.value)
#     #             worklist.append(rowlist)
#     #             if num > 0 and list(sheet.rows).index(i) + 1 == num:
#     #                 break
#     #         # self.isTrue(c,)
#     #         self.logger.info(f"READEXCEL:读取{sheetname}标签页中 行 内容{worklist}成功")
#     #         return worklist
#     #     except:
#     #         self.logger.info(f"READEXCEL:读取{sheetname}标签页中 行 内容失败")
#     #         raise
#
#     def get_all_rows(self, sheetname,num = 0):
#         try:
#             worklist = []
#             sheet = self.workbook[sheetname]
#
#             for i in sheet.rows:
#                 rowlist = []
#                 for j in i:
#                     rowlist.append(j.value)
#                 worklist.append(rowlist)
#             # self.isTrue(c,)
#             self.logger.info(f"READEXCEL:读取{sheetname}标签页中 行 内容{worklist}成功")
#             return worklist
#         except:
#             self.logger.info(f"READEXCEL:读取{sheetname}标签页中 行 内容失败")
#             raise
#
#     # 读取所有列
#     def get_all_cols(self, sheetname):
#         try:
#             collist = []
#             sheet = self.workbook[sheetname]
#             for i in sheet.columns:
#                 rowlist = []
#                 for j in i:
#                     rowlist.append(j.value)
#                 collist.append(rowlist)
#             self.logger.info(f"READEXCEL:读取{sheetname}标签页中 列 内容{collist}成功")
#             return collist
#         except:
#             self.logger.info(f"READEXCEL:读取{sheetname}标签页中 列 内容失败")
#             raise

from log_ri_zhi.logHandler import LogHandler
from openpyxl import load_workbook
from openpyxl import Workbook

logger = LogHandler.logger


# 读
class ReadExcel():
    def __init__(self, path):
        """

        :param path: 文件名
        """
        self.workbook = load_workbook(path)

    # 获取标签页
    def get_sheetnames(self):
        return self.workbook.sheetnames

    # 读取所有行
    def get_all_rows(self, sheetname, num=0, msg="行"):
        """
        读取所有行
        :param sheetname: 标签页名
        :param num: 总共要读num行
        :param msg: "行"表示读出来时行数据 "列"表示读出来是列数据
        :return: 数据list
        """
        try:
            worklist = []
            # 进入标签页
            sheet = self.workbook[sheetname]
            a = sheet.rows
            if msg == "列":
                a = sheet.columns
            for z, i in enumerate(a):
                rowlist = []
                for j in i:
                    rowlist.append(j.value)
                worklist.append(rowlist)
                if num != 0:
                    if num == (z + 1):
                        break
            logger.info(f"READEXCEL:读取{sheetname}标签页中 {msg} 内容{worklist}成功")
            return worklist
        except:
            logger.info(f"READEXCEL:读取{sheetname}标签页中 {msg} 内容失败")
            raise

    # 读取所有列
    def get_all_cols(self, sheetname, num=0):
        return self.get_all_rows(sheetname, num, msg="列")


# 写
class WriteExcel():

    def __init__(self, path, sheetname, mode="w"):
        """
        初始化文件对象
        :param path: 文件路径,名称
        :param sheetname: 标签页名称
        :param mode: 权限 'a'追加 'w'重写
        """
        self.path = path
        if mode == "w":
            self.workbook = Workbook()
            self.sheet = self.workbook.create_sheet(sheetname)
        elif mode == "a":
            self.workbook = load_workbook(self.path)
            self.sheet = self.workbook[sheetname]

    def __del__(self):
        self.workbook.save(self.path)

    def set_all_rows(self, iterable, rows=1):
        """
        写入所有行
        :param iterable: 可迭代对象 eg:[[iter1,iter2,iter3...],[iter1,iter2,iter3...]]
        :param rows: 默认重第一行开始，也可以规定从rows行开始
        :return: None
        """
        for i, j in enumerate(iterable):  # [[1,2,3],[4,5]]
            self.set_line_rows(j, rows + i)

    def set_line_rows(self, iterable, rows=1):
        """
        写入一行
        :param iterable: 可迭代对象 eg:[iter1,iter2,iter3...]
        :param rows: 默认重第一行开始，也可以规定从rows行开始
        :return: None
        """
        try:
            for i, j in enumerate(iterable):
                self.sheet.cell(rows, i + 1, j)
            logger.info(f"WRITEEXCEL:写入{self.path}文件中内容为{iterable}成功")
        except:
            logger.info(f"WRITEEXCEL:写入{self.path}文件中内容{iterable}失败")
            raise
